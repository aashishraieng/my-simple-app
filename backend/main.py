from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from openpyxl import Workbook, load_workbook
from pathlib import Path
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

app = FastAPI()

# CORS config
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Data model
class Registration(BaseModel):
    name: str
    email: str
    password: str

EXCEL_FILE = "registrations.xlsx"

# Create Excel if not exists
if not Path(EXCEL_FILE).exists():
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Email", "Password"])
    wb.save(EXCEL_FILE)

# Function to send email
def send_email(name, email, password):
    sender_email = "shashiashishrai@gmail.com"
    sender_password = "mrgz ojio dden qyre"  # Use Gmail App Password
    receiver_email = "shashiashishrai@gmail.com"  # You can change to another email if needed

    subject = "New Registration Received"
    body = f"""
    A new user has registered:

    Name: {name}
    Email: {email}
    Password: {password}
    """

    # Construct email
    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = receiver_email
    message["Subject"] = subject
    message.attach(MIMEText(body, "plain"))

    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(sender_email, sender_password)
        server.send_message(message)
        server.quit()
        print("Email sent successfully.")
    except Exception as e:
        print("Error sending email:", e)

# API endpoint
@app.post("/register")
async def register_user(data: Registration):
    # Save to Excel
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([data.name, data.email, data.password])
    wb.save(EXCEL_FILE)

    # Send email
    send_email(data.name, data.email, data.password)

    return {"message": "Registration saved and emailed successfully"}
